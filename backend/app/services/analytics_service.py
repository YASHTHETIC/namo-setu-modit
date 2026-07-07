from __future__ import annotations

import math
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any

from sqlalchemy import func, select, and_, extract
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.models.namo_setu import Donation, NamoBooking, Temple
from backend.app.models.modit import (
    Inventory,
    Order,
    OrderItem,
    Organization,
    Product,
    Quotation,
    RFQ,
    Supplier,
)
from backend.app.models.shared import AnalyticsEvent, City, State
from backend.app.models.user import User
from backend.app.models.enums import ProductCode


class AnalyticsService:
    # ── Namo Setu ────────────────────────────────────────────────────────

    @staticmethod
    async def get_namo_analytics(session: AsyncSession, period_days: int = 30) -> dict[str, Any]:
        now = datetime.utcnow()
        cutoff = now - timedelta(days=period_days)
        prev_cutoff = cutoff - timedelta(days=period_days)

        # total bookings
        total_bookings = (
            await session.execute(
                select(func.count(NamoBooking.id)).where(NamoBooking.created_at >= cutoff)
            )
        ).scalar_one()

        prev_bookings = (
            await session.execute(
                select(func.count(NamoBooking.id)).where(
                    and_(NamoBooking.created_at >= prev_cutoff, NamoBooking.created_at < cutoff)
                )
            )
        ).scalar_one()

        booking_growth = ((total_bookings - prev_bookings) / prev_bookings * 100) if prev_bookings else 0.0

        # bookings by status
        by_status_rows = (
            await session.execute(
                select(NamoBooking.booking_status, func.count())
                .where(NamoBooking.created_at >= cutoff)
                .group_by(NamoBooking.booking_status)
            )
        ).all()
        by_status = {row[0]: row[1] for row in by_status_rows}

        # bookings by date
        by_date_rows = (
            await session.execute(
                select(
                    func.date(NamoBooking.created_at).label("day"),
                    func.count(),
                )
                .where(NamoBooking.created_at >= cutoff)
                .group_by("day")
                .order_by("day")
            )
        ).all()
        by_date = {str(row[0]): row[1] for row in by_date_rows}

        # revenue from bookings
        total_booking_revenue = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                        and_(
                            NamoBooking.created_at >= cutoff,
                            NamoBooking.booking_status.in_(["confirmed", "completed"]),
                        )
                    )
                )
            ).scalar_one()
        )

        prev_booking_revenue = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                        and_(
                            NamoBooking.created_at >= prev_cutoff,
                            NamoBooking.created_at < cutoff,
                            NamoBooking.booking_status.in_(["confirmed", "completed"]),
                        )
                    )
                )
            ).scalar_one()
        )

        revenue_growth = (
            ((total_booking_revenue - prev_booking_revenue) / prev_booking_revenue * 100)
            if prev_booking_revenue
            else 0.0
        )

        avg_booking_value = (total_booking_revenue / total_bookings) if total_bookings else 0.0

        # donations
        total_donations = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(Donation.amount), 0)).where(
                        Donation.created_at >= cutoff
                    )
                )
            ).scalar_one()
        )
        prev_donations = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(Donation.amount), 0)).where(
                        and_(Donation.created_at >= prev_cutoff, Donation.created_at < cutoff)
                    )
                )
            ).scalar_one()
        )
        donation_growth = ((total_donations - prev_donations) / prev_donations * 100) if prev_donations else 0.0

        total_revenue = total_booking_revenue + total_donations

        # active temples
        active_temples = (
            await session.execute(
                select(func.count(Temple.id)).where(Temple.is_active == True)
            )
        ).scalar_one()

        # new users
        new_users = (
            await session.execute(
                select(func.count(User.id)).where(User.created_at >= cutoff)
            )
        ).scalar_one()

        # top temples by revenue
        top_temples_rows = (
            await session.execute(
                select(
                    Temple.id,
                    Temple.name,
                    func.coalesce(func.sum(NamoBooking.total_amount), 0).label("rev"),
                    func.count(NamoBooking.id).label("bk_count"),
                )
                .outerjoin(
                    NamoBooking,
                    and_(NamoBooking.temple_id == Temple.id, NamoBooking.created_at >= cutoff),
                )
                .group_by(Temple.id, Temple.name)
                .order_by(func.sum(NamoBooking.total_amount).desc())
                .limit(10)
            )
        ).all()
        top_temples = [
            {
                "temple_id": str(r[0]),
                "temple_name": r[1],
                "revenue": float(r[2]),
                "booking_count": r[3],
            }
            for r in top_temples_rows
        ]

        # bookings by temple
        by_temple_rows = (
            await session.execute(
                select(Temple.name, func.count(NamoBooking.id))
                .join(NamoBooking, NamoBooking.temple_id == Temple.id)
                .where(NamoBooking.created_at >= cutoff)
                .group_by(Temple.name)
                .order_by(func.count(NamoBooking.id).desc())
            )
        ).all()
        bookings_by_temple = {row[0]: row[1] for row in by_temple_rows}

        # conversion rate
        confirmed_count = by_status.get("confirmed", 0) + by_status.get("completed", 0)
        conversion_rate = (confirmed_count / total_bookings * 100) if total_bookings else 0.0

        return {
            "revenue": {"total": total_revenue, "growth_pct": round(revenue_growth, 2)},
            "bookings": {
                "total": total_bookings,
                "growth_pct": round(booking_growth, 2),
                "by_status": by_status,
                "by_date": by_date,
                "by_temple": bookings_by_temple,
            },
            "donations_total": total_donations,
            "donations_growth_pct": round(donation_growth, 2),
            "temples_active": active_temples,
            "users_new": new_users,
            "top_temples": top_temples,
            "avg_booking_value": round(avg_booking_value, 2),
            "conversion_rate": round(conversion_rate, 2),
        }

    # ── MODIT ────────────────────────────────────────────────────────────

    @staticmethod
    async def get_modit_analytics(session: AsyncSession, period_days: int = 30) -> dict[str, Any]:
        now = datetime.utcnow()
        cutoff = now - timedelta(days=period_days)
        prev_cutoff = cutoff - timedelta(days=period_days)

        # orders
        total_orders = (
            await session.execute(
                select(func.count(Order.id)).where(Order.created_at >= cutoff)
            )
        ).scalar_one()
        prev_orders = (
            await session.execute(
                select(func.count(Order.id)).where(
                    and_(Order.created_at >= prev_cutoff, Order.created_at < cutoff)
                )
            )
        ).scalar_one()
        order_growth = ((total_orders - prev_orders) / prev_orders * 100) if prev_orders else 0.0

        # orders by status
        by_status_rows = (
            await session.execute(
                select(Order.status, func.count())
                .where(Order.created_at >= cutoff)
                .group_by(Order.status)
            )
        ).all()
        by_status = {row[0]: row[1] for row in by_status_rows}

        # orders by date
        by_date_rows = (
            await session.execute(
                select(
                    func.date(Order.placed_at).label("day"),
                    func.count(),
                )
                .where(Order.created_at >= cutoff)
                .group_by("day")
                .order_by("day")
            )
        ).all()
        by_date = {str(row[0]): row[1] for row in by_date_rows}

        # order revenue
        total_order_revenue = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(OrderItem.line_total), 0))
                    .join(Order, OrderItem.order_id == Order.id)
                    .where(Order.created_at >= cutoff)
                )
            ).scalar_one()
        )
        prev_order_revenue = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(OrderItem.line_total), 0))
                    .join(Order, OrderItem.order_id == Order.id)
                    .where(and_(Order.created_at >= prev_cutoff, Order.created_at < cutoff))
                )
            ).scalar_one()
        )
        revenue_growth = (
            ((total_order_revenue - prev_order_revenue) / prev_order_revenue * 100)
            if prev_order_revenue
            else 0.0
        )
        avg_order_value = (total_order_revenue / total_orders) if total_orders else 0.0

        # active suppliers
        active_suppliers = (
            await session.execute(
                select(func.count(Supplier.id))
                .join(Organization, Supplier.organization_id == Organization.id)
                .where(Organization.is_active == True)
            )
        ).scalar_one()

        # pending RFQs
        pending_rfqs = (
            await session.execute(
                select(func.count(RFQ.id)).where(RFQ.status.in_(["open", "pending"]))
            )
        ).scalar_one()

        # RFQ conversion
        total_rfq = (
            await session.execute(
                select(func.count(RFQ.id)).where(RFQ.created_at >= cutoff)
            )
        ).scalar_one()
        converted_rfq = (
            await session.execute(
                select(func.count(Quotation.id))
                .join(RFQ, Quotation.rfq_id == RFQ.id)
                .where(RFQ.created_at >= cutoff)
            )
        ).scalar_one()
        rfq_conversion = (converted_rfq / total_rfq * 100) if total_rfq else 0.0

        # inventory value
        inventory_value = float(
            (
                await session.execute(
                    select(func.coalesce(func.sum(Inventory.quantity_on_hand * Product.list_price), 0))
                    .join(Product, Inventory.product_id == Product.id)
                )
            ).scalar_one()
        )

        # top products
        top_products_rows = (
            await session.execute(
                select(
                    Product.id,
                    Product.name,
                    func.coalesce(func.sum(OrderItem.line_total), 0).label("rev"),
                    func.count(OrderItem.id).label("cnt"),
                )
                .join(OrderItem, OrderItem.product_id == Product.id)
                .join(Order, OrderItem.order_id == Order.id)
                .where(Order.created_at >= cutoff)
                .group_by(Product.id, Product.name)
                .order_by(func.sum(OrderItem.line_total).desc())
                .limit(10)
            )
        ).all()
        top_products = [
            {
                "product_id": str(r[0]),
                "product_name": r[1],
                "revenue": float(r[2]),
                "order_count": r[3],
            }
            for r in top_products_rows
        ]

        return {
            "revenue": {"total": total_order_revenue, "growth_pct": round(revenue_growth, 2)},
            "orders": {
                "total": total_orders,
                "growth_pct": round(order_growth, 2),
                "by_status": by_status,
                "by_date": by_date,
            },
            "suppliers_active": active_suppliers,
            "rfqs_pending": pending_rfqs,
            "inventory_value": round(inventory_value, 2),
            "top_products": top_products,
            "avg_order_value": round(avg_order_value, 2),
            "rfq_conversion_rate": round(rfq_conversion, 2),
        }

    # ── Revenue Forecast ─────────────────────────────────────────────────

    @staticmethod
    async def get_revenue_forecast(
        session: AsyncSession, product_code: str, months: int = 6
    ) -> dict[str, Any]:
        now = datetime.utcnow()
        lookback = months * 2

        if product_code == ProductCode.NAMO_SETU:
            monthly_rows = (
                await session.execute(
                    select(
                        extract("year", NamoBooking.created_at).label("yr"),
                        extract("month", NamoBooking.created_at).label("mo"),
                        func.coalesce(func.sum(NamoBooking.total_amount), 0).label("rev"),
                    )
                    .where(NamoBooking.created_at >= now - timedelta(days=lookback * 31))
                    .group_by("yr", "mo")
                    .order_by("yr", "mo")
                )
            ).all()
        else:
            monthly_rows = (
                await session.execute(
                    select(
                        extract("year", Order.placed_at).label("yr"),
                        extract("month", Order.placed_at).label("mo"),
                        func.coalesce(func.sum(OrderItem.line_total), 0).label("rev"),
                    )
                    .join(Order, OrderItem.order_id == Order.id)
                    .where(Order.created_at >= now - timedelta(days=lookback * 31))
                    .group_by("yr", "mo")
                    .order_by("yr", "mo")
                )
            ).all()

        if not monthly_rows:
            return {
                "product_code": product_code,
                "monthly_forecasts": [],
                "trend": "insufficient_data",
                "confidence": 0.0,
            }

        values = [(int(r[0]), int(r[1]), float(r[2])) for r in monthly_rows]
        n = len(values)
        x = list(range(n))
        y = [v[2] for v in values]

        mean_x = sum(x) / n
        mean_y = sum(y) / n
        ss_xy = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, y))
        ss_xx = sum((xi - mean_x) ** 2 for xi in x)

        if ss_xx == 0:
            slope = 0.0
            intercept = mean_y
        else:
            slope = ss_xy / ss_xx
            intercept = mean_y - slope * mean_x

        y_pred = [slope * xi + intercept for xi in x]
        ss_res = sum((yi - yp) ** 2 for yi, yp in zip(y, y_pred))
        ss_tot = sum((yi - mean_y) ** 2 for yi in y)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0.0

        residual_std = math.sqrt(ss_res / (n - 2)) if n > 2 else 0.0

        if slope > mean_y * 0.02:
            trend = "upward"
        elif slope < -mean_y * 0.02:
            trend = "downward"
        else:
            trend = "stable"

        forecasts = []
        last_yr, last_mo = values[-1][0], values[-1][1]
        for i in range(1, months + 1):
            future_x = n + i - 1
            pred = max(slope * future_x + intercept, 0)
            margin = (
                1.96
                * residual_std
                * math.sqrt(1 + 1 / n + (future_x - mean_x) ** 2 / ss_xx)
                if ss_xx > 0
                else 0
            )

            mo = last_mo + i
            yr = last_yr + (mo - 1) // 12
            mo = ((mo - 1) % 12) + 1

            forecasts.append({
                "year": yr,
                "month": mo,
                "forecast": round(pred, 2),
                "lower_bound": round(max(pred - margin, 0), 2),
                "upper_bound": round(pred + margin, 2),
            })

        return {
            "product_code": product_code,
            "monthly_forecasts": forecasts,
            "trend": trend,
            "confidence": round(r_squared * 100, 2),
        }

    # ── Heatmap Data ─────────────────────────────────────────────────────

    @staticmethod
    async def get_heatmap_data(
        session: AsyncSession, product_code: str, entity_type: str = "bookings"
    ) -> dict[str, Any]:
        if product_code == ProductCode.NAMO_SETU and entity_type in ("bookings", "donations"):
            rows = (
                await session.execute(
                    select(
                        City.name.label("city"),
                        State.name.label("state"),
                        City.latitude,
                        City.longitude,
                        func.count().label("count"),
                        func.coalesce(func.sum(NamoBooking.total_amount), 0).label("value"),
                    )
                    .join(Temple, Temple.city_id == City.id)
                    .join(State, Temple.state_id == State.id)
                    .outerjoin(NamoBooking, NamoBooking.temple_id == Temple.id)
                    .group_by(City.name, State.name, City.latitude, City.longitude)
                )
            ).all()
        elif product_code == ProductCode.MODIT:
            rows = (
                await session.execute(
                    select(
                        City.name.label("city"),
                        State.name.label("state"),
                        City.latitude,
                        City.longitude,
                        func.count(Order.id).label("count"),
                        func.coalesce(func.sum(OrderItem.line_total), 0).label("value"),
                    )
                    .join(Organization, Organization.id == Order.organization_id)
                    .join(OrderItem, OrderItem.order_id == Order.id)
                    .join(City, City.id == Organization.id)
                    .join(State, State.id == City.id)
                    .group_by(City.name, State.name, City.latitude, City.longitude)
                )
            ).all()
        else:
            rows = []

        points = [
            {
                "city": r[0],
                "state": r[1],
                "latitude": float(r[2]) if r[2] else None,
                "longitude": float(r[3]) if r[3] else None,
                "count": r[4],
                "value": float(r[5]),
            }
            for r in rows
        ]

        return {
            "points": points,
            "aggregate_by": "city",
            "total_count": sum(p["count"] for p in points),
        }

    # ── Growth Metrics ───────────────────────────────────────────────────

    @staticmethod
    async def get_growth_metrics(
        session: AsyncSession, product_code: str, period_days: int = 30
    ) -> dict[str, Any]:
        now = datetime.utcnow()
        cutoff = now - timedelta(days=period_days)
        prev_cutoff = cutoff - timedelta(days=period_days)

        def _growth(current: int, previous: int) -> dict[str, Any]:
            pct = ((current - previous) / previous * 100) if previous else 0.0
            return {"current": current, "previous": previous, "growth_pct": round(pct, 2)}

        # user growth
        current_users = (
            await session.execute(
                select(func.count(User.id)).where(User.created_at >= cutoff)
            )
        ).scalar_one()
        prev_users = (
            await session.execute(
                select(func.count(User.id)).where(
                    and_(User.created_at >= prev_cutoff, User.created_at < cutoff)
                )
            )
        ).scalar_one()

        result: dict[str, Any] = {
            "period_days": period_days,
            "user_growth": _growth(current_users, prev_users),
        }

        if product_code == ProductCode.NAMO_SETU:
            cur_book = (
                await session.execute(
                    select(func.count(NamoBooking.id)).where(NamoBooking.created_at >= cutoff)
                )
            ).scalar_one()
            prev_book = (
                await session.execute(
                    select(func.count(NamoBooking.id)).where(
                        and_(NamoBooking.created_at >= prev_cutoff, NamoBooking.created_at < cutoff)
                    )
                )
            ).scalar_one()
            cur_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                            and_(
                                NamoBooking.created_at >= cutoff,
                                NamoBooking.booking_status.in_(["confirmed", "completed"]),
                            )
                        )
                    )
                ).scalar_one()
            )
            prev_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                            and_(
                                NamoBooking.created_at >= prev_cutoff,
                                NamoBooking.created_at < cutoff,
                                NamoBooking.booking_status.in_(["confirmed", "completed"]),
                            )
                        )
                    )
                ).scalar_one()
            )
            result["booking_growth"] = _growth(cur_book, prev_book)
            result["revenue_growth"] = {
                "current": cur_rev,
                "previous": prev_rev,
                "growth_pct": round(((cur_rev - prev_rev) / prev_rev * 100) if prev_rev else 0.0, 2),
            }
            result["order_growth"] = {"current": 0, "previous": 0, "growth_pct": 0.0}
        else:
            cur_orders = (
                await session.execute(
                    select(func.count(Order.id)).where(Order.created_at >= cutoff)
                )
            ).scalar_one()
            prev_orders = (
                await session.execute(
                    select(func.count(Order.id)).where(
                        and_(Order.created_at >= prev_cutoff, Order.created_at < cutoff)
                    )
                )
            ).scalar_one()
            cur_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(OrderItem.line_total), 0))
                        .join(Order, OrderItem.order_id == Order.id)
                        .where(Order.created_at >= cutoff)
                    )
                ).scalar_one()
            )
            prev_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(OrderItem.line_total), 0))
                        .join(Order, OrderItem.order_id == Order.id)
                        .where(and_(Order.created_at >= prev_cutoff, Order.created_at < cutoff))
                    )
                ).scalar_one()
            )
            result["order_growth"] = _growth(cur_orders, prev_orders)
            result["revenue_growth"] = {
                "current": cur_rev,
                "previous": prev_rev,
                "growth_pct": round(((cur_rev - prev_rev) / prev_rev * 100) if prev_rev else 0.0, 2),
            }
            result["booking_growth"] = {"current": 0, "previous": 0, "growth_pct": 0.0}

        return result

    # ── AI Insights ──────────────────────────────────────────────────────

    @staticmethod
    async def get_ai_insights(session: AsyncSession, product_code: str) -> dict[str, Any]:
        insights: list[str] = []
        anomalies: list[dict[str, Any]] = []
        recommendations: list[str] = []

        now = datetime.utcnow()
        last_30 = now - timedelta(days=30)
        prev_30 = last_30 - timedelta(days=30)

        if product_code == ProductCode.NAMO_SETU:
            cur_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                            and_(
                                NamoBooking.created_at >= last_30,
                                NamoBooking.booking_status.in_(["confirmed", "completed"]),
                            )
                        )
                    )
                ).scalar_one()
            )
            prev_rev = float(
                (
                    await session.execute(
                        select(func.coalesce(func.sum(NamoBooking.total_amount), 0)).where(
                            and_(
                                NamoBooking.created_at >= prev_30,
                                NamoBooking.created_at < last_30,
                                NamoBooking.booking_status.in_(["confirmed", "completed"]),
                            )
                        )
                    )
                ).scalar_one()
            )

            if prev_rev > 0:
                change = (cur_rev - prev_rev) / prev_rev * 100
                if abs(change) > 25:
                    direction = "increased" if change > 0 else "decreased"
                    anomalies.append({
                        "metric": "revenue",
                        "value": cur_rev,
                        "expected_range": [prev_rev * 0.75, prev_rev * 1.25],
                        "severity": "high" if abs(change) > 50 else "medium",
                        "description": f"Revenue {direction} by {abs(change):.1f}% compared to previous period",
                    })

            low_booked = (
                await session.execute(
                    select(Temple.name, func.count(NamoBooking.id).label("cnt"))
                    .outerjoin(
                        NamoBooking,
                        and_(NamoBooking.temple_id == Temple.id, NamoBooking.created_at >= last_30),
                    )
                    .where(Temple.is_active == True)
                    .group_by(Temple.name)
                    .having(func.count(NamoBooking.id) < 5)
                )
            ).all()
            if low_booked:
                names = [r[0] for r in low_booked[:5]]
                insights.append(
                    f"{len(low_booked)} temples have fewer than 5 bookings in the last 30 days: {', '.join(names)}"
                )
                recommendations.append("Consider promotional campaigns for low-traffic temples")

            recent = (
                await session.execute(
                    select(func.count(NamoBooking.id)).where(NamoBooking.created_at >= last_30)
                )
            ).scalar_one()
            if recent == 0:
                insights.append("No bookings recorded in the last 30 days")
                recommendations.append("Review booking flow and marketing channels")

        else:  # MODIT
            cur_orders = (
                await session.execute(
                    select(func.count(Order.id)).where(Order.created_at >= last_30)
                )
            ).scalar_one()
            prev_orders = (
                await session.execute(
                    select(func.count(Order.id)).where(
                        and_(Order.created_at >= prev_30, Order.created_at < last_30)
                    )
                )
            ).scalar_one()

            if prev_orders > 0:
                change = (cur_orders - prev_orders) / prev_orders * 100
                if abs(change) > 30:
                    direction = "increased" if change > 0 else "decreased"
                    anomalies.append({
                        "metric": "order_volume",
                        "value": cur_orders,
                        "expected_range": [int(prev_orders * 0.7), int(prev_orders * 1.3)],
                        "severity": "high" if abs(change) > 60 else "medium",
                        "description": f"Order volume {direction} by {abs(change):.1f}%",
                    })

            low_stock = (
                await session.execute(
                    select(func.count(Inventory.id)).where(
                        Inventory.quantity_on_hand <= Inventory.reorder_level
                    )
                )
            ).scalar_one()
            if low_stock:
                insights.append(f"{low_stock} products are at or below reorder level")
                recommendations.append("Review inventory and trigger restocking for low-stock items")

            pending = (
                await session.execute(
                    select(func.count(RFQ.id)).where(RFQ.status.in_(["open", "pending"]))
                )
            ).scalar_one()
            if pending > 10:
                insights.append(f"{pending} RFQs are still pending — consider follow-up with suppliers")
                recommendations.append("Prioritize aging RFQs to maintain supplier relationships")

            active_suppliers = (
                await session.execute(
                    select(func.count(func.distinct(Supplier.id)))
                    .join(Organization, Supplier.organization_id == Organization.id)
                    .where(Organization.is_active == True)
                )
            ).scalar_one()
            if active_suppliers < 5:
                insights.append(f"Only {active_suppliers} active suppliers — limited procurement options")
                recommendations.append("Onboard additional suppliers to diversify sourcing")

        event_count = (
            await session.execute(
                select(func.count(AnalyticsEvent.id)).where(AnalyticsEvent.created_at >= last_30)
            )
        ).scalar_one()
        insights.append(f"{event_count} analytics events tracked in the last 30 days")

        if not insights:
            insights.append("System operating within normal parameters")
        if not recommendations:
            recommendations.append("Continue monitoring key metrics for trend changes")

        return {
            "product_code": product_code,
            "insights": insights,
            "anomalies": anomalies,
            "recommendations": recommendations,
        }

    # ── PDF Export ───────────────────────────────────────────────────────

    @staticmethod
    def export_pdf(analytics_data: dict[str, Any], title: str = "Analytics Report") -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            return b"reportlab not installed"

        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        elements: list = []

        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"])
        )
        elements.append(Spacer(1, 20))

        revenue = analytics_data.get("revenue", {})
        if revenue:
            elements.append(Paragraph("Revenue Overview", styles["Heading2"]))
            rev_data = [
                ["Metric", "Value"],
                ["Total Revenue", f"\u20b9{revenue.get('total', 0):,.2f}"],
                ["Growth", f"{revenue.get('growth_pct', 0):.1f}%"],
            ]
            t = Table(rev_data, colWidths=[3 * inch, 3 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a90d9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 15))

        bookings = analytics_data.get("bookings") or analytics_data.get("orders")
        if bookings:
            label = "Booking" if "by_temple" in bookings else "Order"
            elements.append(Paragraph(f"{label} Overview", styles["Heading2"]))
            bk_data = [
                ["Metric", "Value"],
                [f"Total {label}s", str(bookings.get("total", 0))],
                ["Growth", f"{bookings.get('growth_pct', 0):.1f}%"],
            ]
            for status, count in bookings.get("by_status", {}).items():
                bk_data.append([f"  {status}", str(count)])
            t = Table(bk_data, colWidths=[3 * inch, 3 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#27ae60")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 15))

        top = analytics_data.get("top_temples") or analytics_data.get("top_products")
        if top:
            item_label = "Temple" if "temple_name" in (top[0] if top else {}) else "Product"
            elements.append(Paragraph(f"Top {item_label}s by Revenue", styles["Heading2"]))
            top_data = [[item_label, "Revenue", "Count"]]
            for item in top:
                name = item.get("temple_name") or item.get("product_name", "")
                top_data.append([
                    name,
                    f"\u20b9{item.get('revenue', 0):,.2f}",
                    str(item.get("booking_count", item.get("order_count", 0))),
                ])
            t = Table(top_data, colWidths=[2.5 * inch, 2 * inch, 1.5 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8e44ad")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 15))

        insights = analytics_data.get("insights", [])
        if insights:
            elements.append(Paragraph("Insights", styles["Heading2"]))
            for ins in insights:
                elements.append(Paragraph(f"\u2022 {ins}", styles["Normal"]))
            elements.append(Spacer(1, 15))

        forecast = revenue.get("forecast")
        if forecast:
            elements.append(Paragraph("Revenue Forecast", styles["Heading2"]))
            fc_data = [["Month", "Forecast", "Lower", "Upper"]]
            for fc in forecast:
                fc_data.append([
                    f"{fc.get('year')}-{fc.get('month'):02d}",
                    f"\u20b9{fc.get('forecast', 0):,.2f}",
                    f"\u20b9{fc.get('lower_bound', 0):,.2f}",
                    f"\u20b9{fc.get('upper_bound', 0):,.2f}",
                ])
            t = Table(fc_data, colWidths=[1.5 * inch, 1.8 * inch, 1.5 * inch, 1.5 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e67e22")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]))
            elements.append(t)

        doc.build(elements)
        return buffer.getvalue()

    # ── Excel Export ─────────────────────────────────────────────────────

    @staticmethod
    def export_excel(analytics_data: dict[str, Any], title: str = "Analytics Report") -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, LineChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return b"openpyxl not installed"

        from io import BytesIO

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def _style_header(ws, row: int, cols: int):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # ── Summary sheet ──
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

        row = 4
        revenue = analytics_data.get("revenue", {})
        if revenue:
            ws.cell(row=row, column=1, value="Revenue Overview").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Total Revenue")
            ws.cell(row=row, column=2, value=revenue.get("total", 0))
            ws.cell(row=row, column=2).number_format = '\u20b9#,##0.00'
            row += 1
            ws.cell(row=row, column=1, value="Growth %")
            ws.cell(row=row, column=2, value=f"{revenue.get('growth_pct', 0):.1f}%")
            row += 2

        bookings = analytics_data.get("bookings") or analytics_data.get("orders")
        if bookings:
            label = "Bookings" if "by_temple" in bookings else "Orders"
            ws.cell(row=row, column=1, value=f"{label} Overview").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value=f"Total {label}")
            ws.cell(row=row, column=2, value=bookings.get("total", 0))
            row += 1
            ws.cell(row=row, column=1, value="Growth %")
            ws.cell(row=row, column=2, value=f"{bookings.get('growth_pct', 0):.1f}%")
            row += 2

            by_status = bookings.get("by_status", {})
            if by_status:
                ws.cell(row=row, column=1, value="By Status").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value="Status")
                ws.cell(row=row, column=2, value="Count")
                _style_header(ws, row, 2)
                row += 1
                for status, count in by_status.items():
                    ws.cell(row=row, column=1, value=status)
                    ws.cell(row=row, column=2, value=count)
                    ws.cell(row=row, column=1).border = border
                    ws.cell(row=row, column=2).border = border
                    row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30

        # ── Daily Data sheet ──
        by_date = {}
        if bookings:
            by_date = bookings.get("by_date", {})
        if by_date:
            ws2 = wb.create_sheet("Daily Data")
            ws2.cell(row=1, column=1, value="Daily Counts").font = Font(bold=True, size=12)
            ws2.cell(row=2, column=1, value="Date")
            ws2.cell(row=2, column=2, value="Count")
            _style_header(ws2, 2, 2)
            r = 3
            for dt, cnt in by_date.items():
                ws2.cell(row=r, column=1, value=dt)
                ws2.cell(row=r, column=2, value=cnt)
                ws2.cell(row=r, column=1).border = border
                ws2.cell(row=r, column=2).border = border
                r += 1
            ws2.column_dimensions["A"].width = 18
            ws2.column_dimensions["B"].width = 15
            if r > 3:
                chart = LineChart()
                chart.title = "Daily Trend"
                chart.y_axis.title = "Count"
                data_ref = Reference(ws2, min_col=2, min_row=2, max_row=r - 1)
                cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=r - 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 25
                chart.height = 15
                ws2.add_chart(chart, "D2")

        # ── Top Items sheet ──
        top = analytics_data.get("top_temples") or analytics_data.get("top_products")
        if top:
            ws3 = wb.create_sheet("Top Items")
            item_label = "Temple" if "temple_name" in (top[0] if top else {}) else "Product"
            ws3.cell(row=1, column=1, value=f"Top {item_label}s").font = Font(bold=True, size=12)
            ws3.cell(row=2, column=1, value=item_label)
            ws3.cell(row=2, column=2, value="Revenue")
            ws3.cell(row=2, column=3, value="Count")
            _style_header(ws3, 2, 3)
            r = 3
            for item in top:
                name = item.get("temple_name") or item.get("product_name", "")
                ws3.cell(row=r, column=1, value=name)
                ws3.cell(row=r, column=2, value=item.get("revenue", 0))
                ws3.cell(row=r, column=2).number_format = '\u20b9#,##0.00'
                ws3.cell(row=r, column=3, value=item.get("booking_count", item.get("order_count", 0)))
                for c in range(1, 4):
                    ws3.cell(row=r, column=c).border = border
                r += 1
            ws3.column_dimensions["A"].width = 30
            ws3.column_dimensions["B"].width = 18
            ws3.column_dimensions["C"].width = 12
            if r > 3:
                chart = BarChart()
                chart.title = f"Top {item_label}s by Revenue"
                chart.y_axis.title = "Revenue (\u20b9)"
                data_ref = Reference(ws3, min_col=2, min_row=2, max_row=r - 1)
                cats_ref = Reference(ws3, min_col=1, min_row=3, max_row=r - 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 25
                chart.height = 15
                ws3.add_chart(chart, "E2")

        # ── Forecast sheet ──
        forecast_list = revenue.get("forecast") if revenue else None
        if forecast_list:
            ws4 = wb.create_sheet("Forecast")
            ws4.cell(row=1, column=1, value="Revenue Forecast").font = Font(bold=True, size=12)
            ws4.cell(row=2, column=1, value="Month")
            ws4.cell(row=2, column=2, value="Forecast")
            ws4.cell(row=2, column=3, value="Lower Bound")
            ws4.cell(row=2, column=4, value="Upper Bound")
            _style_header(ws4, 2, 4)
            r = 3
            for fc in forecast_list:
                ws4.cell(row=r, column=1, value=f"{fc.get('year')}-{fc.get('month'):02d}")
                ws4.cell(row=r, column=2, value=fc.get("forecast", 0))
                ws4.cell(row=r, column=3, value=fc.get("lower_bound", 0))
                ws4.cell(row=r, column=4, value=fc.get("upper_bound", 0))
                for c in range(1, 5):
                    ws4.cell(row=r, column=c).border = border
                    if c > 1:
                        ws4.cell(row=r, column=c).number_format = '\u20b9#,##0.00'
                r += 1
            if r > 3:
                chart = LineChart()
                chart.title = "Revenue Forecast"
                chart.y_axis.title = "Revenue (\u20b9)"
                for col_idx in [2, 3, 4]:
                    data_ref = Reference(ws4, min_col=col_idx, min_row=2, max_row=r - 1)
                    chart.add_data(data_ref, titles_from_data=True)
                cats_ref = Reference(ws4, min_col=1, min_row=3, max_row=r - 1)
                chart.set_categories(cats_ref)
                chart.width = 25
                chart.height = 15
                ws4.add_chart(chart, "F2")

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
